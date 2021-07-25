import streamlit as st
import os
from openpyxl import load_workbook

def show_survey_questions() :
    filename="Survey.xlsx"
    userform = st.form(key='user-form')
    email=userform.text_input(label='Enter WIBD email')
    kryptonite = userform.selectbox('What is your kryptonite?',('Procrastination', 'Perfectionism', 'Self-doubt','Lack of focus','Lack of follow-through'))
    success = userform.selectbox('How do you define success?',('Flying first class', 'Finding my purpose in life', 'Having a family that loves me'))
    energy = userform.selectbox('When do you have the best energy?',('6am-12pm', '12pm-6pm', '6pm-Midnight','Midnight-6am'))
    socialmedia = userform.selectbox('What is your favorite social media network?',('Twitter','Facebook','Instagram','Snapchat','LinkedIn'))
    submit_button = userform.form_submit_button(label='Submit')

    if(submit_button) :
        dirname = os.path.dirname(__file__)
        filepath= os.path.join(dirname, filename)
        st.write(dirname)
        st.write(filepath)
        anwsers=[email,kryptonite,success,energy,socialmedia]
        wb = load_workbook(filepath)
        sheet = wb.worksheets[0]
        row = sheet.max_row+1
        col=1;
        for ans in anwsers:
            sheet.cell(row,col).value=ans;
            col=col+1;
            wb.save(filename)
        wb = load_workbook(filepath)
        sheet = wb.worksheets[0]
        max_row = sheet.max_row + 1
        max_col=sheet.max_column+1
        for i in range(1, max_row):
            for j in range(1,max_col):
                cell_obj = sheet.cell(row=i, column=j)
                st.write(f' {i},{j} : {cell_obj.value} ')
        wb.save(filename)




