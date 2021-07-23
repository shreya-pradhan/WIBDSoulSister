import streamlit as st
from openpyxl import load_workbook
filename="Survey.xlsx"
form = st.form(key='my_form')
email=form.text_input(label='Enter WIBD email')
kryptonite = form.selectbox('What is your kryptonite?',('Procrastination', 'Perfectionism', 'Self-doubt','Lack of focus','Lack of follow-through'))
success = form.selectbox('How do you define success?',('Flying first class', 'Finding my purpose in life', 'Having a family that loves me'))
energy = form.selectbox('When do you have the best energy?',('6am-12pm', '12pm-6pm', '6pm-Midnight','Midnight-6am'))
socialmedia = form.selectbox('What is your favorite social media network?',('Twitter','Facebook','Instagram','Snapchat','LinkedIn'))
submit_button = form.form_submit_button(label='Submit')

if(submit_button) :
    anwsers=[email,kryptonite,success,energy,socialmedia]
    wb = load_workbook(filename)
    sheet = wb.worksheets[0]
    row = sheet.max_row+1
    col=1;
    for ans in anwsers:
        st.write(f' {ans} {row}{col} ')
        sheet.cell(row,col).value=ans;
        col=col+1;
    wb.save(filename)
    st.write(f' {row} rows in sheet ')
    st.write(f'hello {email} your information has been saved')
    st.write(f' {socialmedia}  {kryptonite} {success} {energy}')



