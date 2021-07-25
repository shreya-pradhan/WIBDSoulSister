import streamlit as st
import os
from openpyxl import load_workbook



filename = "Survey.xlsx"
dirname = os.path.dirname(__file__)
filepath = os.path.join(dirname, filename)

def write_local_file(answers):

    st.write(dirname)
    st.write(filepath)
    wb = load_workbook(filepath)
    sheet = wb.worksheets[0]
    row = sheet.max_row + 1
    col = 1;
    for ans in answers:
        sheet.cell(row, col).value = ans;
        col = col + 1;
        wb.save(filename)

def read_local_files():
    wb = load_workbook(filepath)
    sheet = wb.worksheets[0]
    max_row = sheet.max_row + 1
    max_col = sheet.max_column + 1
    for i in range(1, max_row):
        for j in range(1, max_col):
            cell_obj = sheet.cell(row=i, column=j)
            st.write(f' {i},{j} : {cell_obj.value} ')
    wb.save(filename)

